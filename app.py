import os
import uuid
import json
import base64
import io
from pathlib import Path
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from html import escape

import streamlit as st
from groq import Groq, RateLimitError
from langchain_groq import ChatGroq
from langchain_core.messages import HumanMessage, AIMessage, SystemMessage
from supabase import create_client, Client


# ============================================================
# PAGE CONFIG
# ============================================================

st.set_page_config(
    page_title="RacharlaGPT",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ============================================================
# APP CONFIG
# ============================================================

APP_NAME = "RacharlaGPT"
APP_URL = "https://racharlagpt.streamlit.app/"
CREATOR_NAME = "Racharla Saikrishna"
SUPPORTERS_NAME = "RSKT"
CREATOR_CREDIT = f"Created and developed by {CREATOR_NAME}"

PRIMARY_MODEL = "openai/gpt-oss-20b"
BACKUP_MODEL = "llama-3.1-8b-instant"
VISION_MODEL = "qwen/qwen3.6-27b"
TRANSCRIPTION_MODEL = "whisper-large-v3-turbo"

MAX_CONTEXT_MESSAGES = 14

MODEL_OPTIONS = [
    "Auto (recommended)",
    PRIMARY_MODEL,
    BACKUP_MODEL,
]

# ManaTechSaavy course intelligence. These choices only guide the AI;
# they do not change your existing chat database, authentication, or UI.
COURSE_OPTIONS = [
    "Auto Detect",
    "Python",
    "SQL",
    "Excel",
    "Power BI",
    "JavaScript",
    "Coding",
    "Data Analytics",
]

LEARNING_MODES = [
    "Auto (best for the question)",
    "Learn — understand the concept",
    "Practice — exercises with hints",
    "Build — real-world project",
    "Interview — job preparation",
    "Debug — find and fix mistakes",
]

LEVELS = [
    "Auto Detect",
    "Beginner",
    "Intermediate",
    "Advanced",
]

ANSWER_STYLES = [
    "Balanced",
    "Simple and step-by-step",
    "Professional and concise",
    "Detailed with examples",
]


COURSE_INTELLIGENCE_PROMPT = """
You are the dedicated practical learning instructor inside RacharlaGPT for ManaTechSaavy.
The core learning areas are Python, SQL, Excel, Power BI, JavaScript, Coding, Data Analytics, and related programming/technology topics. Treat these as an expanding teaching-assistant scope rather than a closed list.
The teaching philosophy is: LEARN -> PRACTICE -> BUILD -> GROW.
The audience includes students, beginners changing careers, job seekers, and working professionals.

COURSE TEACHING RULES:
1. Identify the user's intent and teach at the selected level without making the student feel judged.
2. Prefer practical, job-relevant examples over abstract theory.
3. For technical questions, explain the idea first, then show a correct example, then give a small practice task when useful.
4. For code, use clean runnable examples, explain important lines, and point out common mistakes.
5. For SQL, use realistic tables/data and show the query plus expected result when useful.
6. For Excel, give the exact formula/function and explain where to place it; mention common mistakes such as wrong ranges or cell references.
7. For Power BI, distinguish Power Query, data modeling, DAX, and report/visual steps clearly when relevant.
8. For JavaScript, prefer modern, readable examples and explain browser/DOM concepts when relevant.
9. For Data Analytics, connect the workflow across Excel, SQL, Python, and Power BI when that helps.
10. For interview questions, give the answer, why it is correct, and a short interview-ready response.
11. For practice mode, do not immediately reveal the full answer unless the learner asks; start with a hint or guided steps.
12. For build mode, propose realistic mini-projects with requirements, steps, validation, and extension ideas.
13. For debug mode, identify the error, explain the cause, provide the corrected version, and show how to prevent it.
14. For other programming, software, technology, computer science, AI, databases, statistics, career, interview, productivity, education, or general questions, answer helpfully even when they are outside the named core courses. Do not force every question into a course.
15. For medical/MBBS study requests, act as an educational study assistant. Do not provide pirated or unauthorized copyrighted textbook PDFs or links. Instead, offer legal/open-access resources when appropriate, explain topics, create notes, summaries, flashcards, MCQs, study plans, and help analyze user-provided material. For diagnosis or treatment questions, clearly distinguish educational information from professional medical advice.
16. Understand Indian English, common learner wording, typos, and phonetic spelling.
17. Never invent facts, functions, syntax, or tool behavior. If something depends on a version, say so.
18. Use headings, numbered steps, tables, code blocks, examples, and checklists when they improve learning.
19. Keep simple answers simple; increase depth when the learner asks for detail.
"""


DEFAULT_SYSTEM_PROMPT = """
You are RacharlaGPT, a helpful, intelligent, friendly and practical AI assistant.

IMPORTANT BEHAVIOR:

1. Answer the user's actual question directly.
2. If the user makes an obvious spelling mistake, typo, or phonetic mistake,
   understand the intended meaning and answer instead of stopping.
3. If a word is ambiguous and could have multiple meanings, briefly explain
   your interpretation and continue helping.
4. When the user asks "today", "tomorrow", "yesterday", "what day is it",
   "day of the week", "date today", or similar questions, use the current
   date and weekday supplied in the system context.
5. Never invent a special day, holiday, festival, observance, anniversary,
   or historical event. If the user asks about a "speciality" of today and
   there is no reliable information available in the supplied context,
   clearly say that there is no confirmed major observance rather than
   making something up.
6. If the user asks about today's date/day, give the exact date and weekday.
7. Use headings, bullet points, numbered steps and tables when useful.
8. If the user asks for code, provide complete working code.
9. Explain important code changes briefly.
10. Be concise for simple questions and detailed when the user needs detail.
11. If something is uncertain, say so clearly.
12. Do not stop answering merely because the user's spelling is imperfect.
13. Understand Indian English and common conversational spelling variations.
14. Be practical and solution-oriented.
15. RacharlaGPT was created and developed by Racharla Saikrishna, with supporters RSKT.
16. If a user asks who created, developed, built, or made RacharlaGPT, answer clearly: Racharla Saikrishna. If they ask about supporters, say: RSKT.
17. Do not claim a different creator or developer.
18. When a user asks to analyze, review, improve, optimize, refactor, fix, or "make this better", use this workflow when applicable: ANALYZE -> IDENTIFY ISSUES -> IMPROVE -> SHOW THE BETTER VERSION -> EXPLAIN WHY -> SUGGEST NEXT STEPS. For code, preserve working behavior unless the user asks for a redesign.
19. If the user says "analyze and make better" or similar wording, treat it as an explicit request for both diagnosis and an improved result, not just general advice.
20. If the user attaches an image, screenshot, screen chat, diagram, chart, or photo, use the attachment context to answer the user's question. Do not ignore the attachment.

21. Preserve the strongest existing teaching-assistant behavior. Do not reduce teaching quality when adding broader assistant capabilities.
22. Educational answers should be readable, structured, and student-friendly. Prefer: Direct answer -> Key idea -> Step-by-step explanation -> Example -> Common mistakes -> Quick recap -> Practice/next step, when appropriate.
23. For Python, SQL, Excel, Power BI, JavaScript, coding, data analytics, AI, databases, statistics and related technical topics, explain at the learner's level, use clean examples, explain important code lines, and show expected output when useful.
24. For notes and study requests, use clear headings, numbered steps, bullets, concise definitions, examples, memory tips, and exam/interview points. Avoid large walls of text.
25. Keep simple questions simple, but provide enough explanation to make the answer understandable and actionable.
26. When the user asks to "analyze and make better", use ANALYZE -> IDENTIFY ISSUES -> IMPROVE -> SHOW BETTER VERSION -> EXPLAIN WHY -> NEXT STEPS.
27. Continue answering general questions normally. Do not force general questions into programming or courses.
28. Keep the interface welcoming and student-focused. After successful login, greet the user naturally using their available display name when present.
29. If an authenticated profile provides a safe public avatar/photo URL, the UI may display it as the user's profile image. Never expose private tokens or sensitive account information.
30. Allow users to optionally upload a profile photo and use it as a local profile avatar. Do not require a photo.
31. Add a concise Getting Started area for new users explaining Learn, Practice, Build, Analyze, Upload, Voice, and General Assistant capabilities.
32. Keep the UI clean and readable: clear hierarchy, compact cards, consistent spacing, accessible contrast, and no excessive decoration.
33. Preserve all existing teaching-assistant readability, structured notes, course intelligence, authentication, chat history, exports, image/document handling, and voice capabilities while improving the UI.

21. If the user attaches a document or data file, use its extracted contents as source context when answering. Clearly distinguish what comes from the attachment from general knowledge.
22. For voice input, treat the transcription as the user's actual question. Do not complain about spelling or transcription imperfections; infer obvious speech-recognition errors from context.
23. Never claim to have seen or read an attachment when no usable attachment context is available.
"""


# ============================================================
# FILE PATHS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent


def resolve_asset(filename, keywords=()):
    """Find an asset reliably across normal repo asset locations and legacy copies."""
    candidates = [
        BASE_DIR / "assets" / filename,
        BASE_DIR / filename,
        Path.cwd() / "assets" / filename,
        Path.cwd() / filename,
        Path("/mnt/data/assets") / filename,
    ]
    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate

    normalized = {str(k).lower().replace("-", "_").replace(" ", "_") for k in keywords}
    for root in [BASE_DIR, Path.cwd(), Path("/mnt/data/assets")]:
        if not root.exists():
            continue
        try:
            for candidate in root.rglob("*"):
                if not candidate.is_file():
                    continue
                name = candidate.name.lower().replace("-", "_").replace(" ", "_")
                if normalized and all(k in name for k in normalized):
                    return candidate
        except Exception:
            continue
    return None


LOGO_PATH = resolve_asset("manatechsaavy_logo.png", ("manatechsaavy", "logo"))
BANNER_PATH = resolve_asset("manatechsaavy_banner.jfif", ("manatechsaavy", "banner"))


# ============================================================
# SECRETS
# ============================================================

def get_secret(name):
    value = os.environ.get(name)

    if value:
        return value.strip()

    try:
        value = st.secrets.get(name)

        if value:
            return str(value).strip()

    except Exception:
        pass

    return None


GROQ_API_KEY = get_secret("GROQ_API_KEY")
SUPABASE_URL = get_secret("SUPABASE_URL")
SUPABASE_KEY = get_secret("SUPABASE_KEY")


# ============================================================
# CONFIGURATION CHECK
# ============================================================

if not SUPABASE_URL or not SUPABASE_KEY:
    st.error("Supabase is not configured.")

    st.info(
        "Add SUPABASE_URL and SUPABASE_KEY in "
        "Streamlit Cloud → Manage app → Settings → Secrets."
    )

    st.stop()


if not GROQ_API_KEY:
    st.error("GROQ_API_KEY is not configured.")
    st.stop()


os.environ["GROQ_API_KEY"] = GROQ_API_KEY


# ============================================================
# SUPABASE CLIENT
# ============================================================

if "supabase_client" not in st.session_state:
    # IMPORTANT:
    # Do not pass ClientOptions here. Some Streamlit Cloud environments may
    # install supabase-py 2.24.x, where ClientOptions has a known regression:
    # client creation can fail with:
    # AttributeError: 'ClientOptions' object has no attribute 'storage'
    #
    # The plain create_client() path uses the library defaults and works
    # across the affected versions. Keep the same client in session state so
    # the OAuth flow can continue when Streamlit returns from Google.
    st.session_state.supabase_client = create_client(
        SUPABASE_URL,
        SUPABASE_KEY,
    )

supabase: Client = st.session_state.supabase_client


# ============================================================
# INDIA DATE / TIME CONTEXT
# ============================================================

def current_india_datetime():
    try:
        return datetime.now(ZoneInfo("Asia/Kolkata"))
    except Exception:
        return datetime.now(timezone.utc)


def date_context():
    now = current_india_datetime()

    return (
        f"Current date: {now.strftime('%d %B %Y')}\n"
        f"Current day of the week: {now.strftime('%A')}\n"
        f"Current time in India (IST): {now.strftime('%I:%M %p')}\n"
        f"Timezone: Asia/Kolkata (IST)"
    )


# ============================================================
# CSS — UI ONLY
# ============================================================

st.markdown(
    """
    <style>

    /* ========================================================
       GLOBAL
       ======================================================== */

    .stApp {
        background: #ffffff;
    }

    .block-container {
        max-width: 1280px;
        padding-top: 1.1rem;
        padding-bottom: 1rem;
    }

    section[data-testid="stSidebar"] {
        background: #f7f8fc;
        border-right: 1px solid #e7e9ef;
    }

    section[data-testid="stSidebar"] .block-container {
        padding-top: 1rem;
        padding-left: 0.9rem;
        padding-right: 0.9rem;
    }


    /* ========================================================
       AUTH PAGE
       ======================================================== */

    .auth-left {
        padding: 10px 18px 8px 4px;
    }

    .auth-logo-wrap {
        margin-bottom: 7px;
    }

    .auth-heading {
        font-size: 32px;
        font-weight: 800;
        letter-spacing: -1px;
        color: #172033;
        margin: 4px 0 4px 0;
    }

    .auth-description {
        color: #687386;
        font-size: 15px;
        line-height: 1.5;
        margin-bottom: 12px;
        max-width: 600px;
    }

    .auth-badge-row {
        display: flex;
        flex-wrap: wrap;
        gap: 7px;
        margin-bottom: 14px;
    }

    .auth-badge {
        display: inline-block;
        padding: 5px 10px;
        border-radius: 999px;
        background: #eef4ff;
        color: #2459b8;
        font-size: 12px;
        font-weight: 700;
    }

    .auth-banner-box {
        width: 100%;
        max-width: 590px;
        margin: 6px auto 0 auto;
    }

    .auth-right {
        padding: 5px 4px 0 4px;
    }

    .auth-card-title {
        font-size: 29px;
        font-weight: 800;
        color: #172033;
        margin-bottom: 3px;
    }

    .auth-card-subtitle {
        color: #718096;
        font-size: 14px;
        line-height: 1.45;
        margin-bottom: 13px;
    }

    div[data-testid="stForm"] {
        border: 1px solid #e1e5ec;
        border-radius: 15px;
        padding: 18px 18px 14px 18px;
        background: #ffffff;
        box-shadow: 0 7px 25px rgba(20, 35, 70, 0.06);
    }

    div[data-testid="stForm"] label {
        font-weight: 600;
        color: #344054;
    }

    div[data-testid="stForm"] input {
        border-radius: 10px;
    }

    button[kind="primary"] {
        border-radius: 10px !important;
        font-weight: 700 !important;
        min-height: 43px !important;
    }

    div[data-baseweb="tab-list"] {
        gap: 4px;
        border-bottom: 1px solid #e5e7eb;
        margin-bottom: 10px;
    }

    button[data-baseweb="tab"] {
        font-weight: 600;
        padding-left: 9px;
        padding-right: 9px;
    }

    /* Google SSO Divider */
    .divider-container {
        display: flex;
        align-items: center;
        text-align: center;
        margin: 15px 0;
        color: #8a94a6;
        font-size: 12px;
    }
    .divider-container::before, .divider-container::after {
        content: '';
        flex: 1;
        border-bottom: 1px solid #e1e5ec;
    }
    .divider-container:not(:empty)::before {
        margin-right: .5em;
    }
    .divider-container:not(:empty)::after {
        margin-left: .5em;
    }


    /* ========================================================
       MAIN BRANDING
       ======================================================== */

    .main-brand-name {
        font-size: 29px;
        font-weight: 800;
        color: #172033;
        letter-spacing: -0.8px;
        line-height: 1.1;
        margin-top: 3px;
    }

    .main-brand-subtitle {
        color: #7a8494;
        font-size: 13px;
        margin-top: 3px;
    }


    /* ========================================================
       WELCOME
       ======================================================== */

    .welcome-box {
        max-width: 780px;
        margin: 8vh auto 4vh auto;
        text-align: center;
        padding: 22px 16px;
    }

    .welcome-title {
        font-size: 31px;
        font-weight: 800;
        color: #172033;
        letter-spacing: -0.7px;
        margin-bottom: 8px;
    }

    .welcome-text {
        color: #707b8c;
        line-height: 1.65;
        max-width: 680px;
        margin: 0 auto;
        font-size: 15px;
    }


    /* ========================================================
       CHAT
       ======================================================== */

    div[data-testid="stChatMessage"] {
        border-radius: 14px;
        padding-top: 6px;
        padding-bottom: 6px;
    }

    .stChatInputContainer {
        padding-bottom: 7px;
    }


    /* ========================================================
       SIDEBAR
       ======================================================== */

    section[data-testid="stSidebar"] button {
        border-radius: 99px;
    }

    section[data-testid="stSidebar"] .stButton button {
        min-height: 37px;
    }

    section[data-testid="stSidebar"] .stDownloadButton button {
        border-radius: 99px;
        min-height: 37px;
    }

    section[data-testid="stSidebar"] hr {
        margin-top: 10px;
        margin-bottom: 10px;
    }


    /* ========================================================
       MOBILE
       ======================================================== */

    @media (max-width: 768px) {

        .block-container {
            padding-left: 0.75rem;
            padding-right: 0.75rem;
            padding-top: 0.6rem;
        }

        .auth-left {
            padding: 2px 2px 8px 2px;
            text-align: center;
        }

        .auth-logo-wrap {
            margin-bottom: 3px;
        }

        .auth-heading {
            font-size: 26px;
        }

        .auth-description {
            font-size: 14px;
            margin: 0 auto 9px auto;
        }

        .auth-badge-row {
            justify-content: center;
            margin-bottom: 9px;
        }

        .auth-banner-box {
            max-width: 100%;
            margin-top: 4px;
        }

        .auth-right {
            padding: 3px 0 0 0;
        }

        .auth-card-title {
            font-size: 25px;
        }

        .auth-card-subtitle {
            font-size: 13px;
        }

        div[data-testid="stForm"] {
            padding: 15px 13px 11px 13px;
            border-radius: 13px;
            box-shadow: none;
        }

        .main-brand-name {
            font-size: 25px;
        }

        .main-brand-subtitle {
            font-size: 12px;
        }

        .welcome-box {
            margin: 6vh auto 4vh auto;
            padding: 15px 7px;
        }

        .welcome-title {
            font-size: 25px;
        }

        .welcome-text {
            font-size: 14px;
        }
    }


    /* ========================================================
       SMALL PHONES
       ======================================================== */

    @media (max-width: 430px) {

        .auth-heading {
            font-size: 24px;
        }

        .auth-description {
            font-size: 13px;
        }

        .auth-card-title {
            font-size: 23px;
        }

        .main-brand-name {
            font-size: 23px;
        }

        .main-brand-subtitle {
            font-size: 11px;
        }

        .welcome-title {
            font-size: 23px;
        }
    }


    .rgpt-welcome {
        padding: 0.4rem 0 0.9rem 0;
    }

    .rgpt-kicker {
        font-size: 0.72rem;
        font-weight: 700;
        letter-spacing: 0.12em;
        opacity: 0.7;
    }

    .rgpt-welcome h1 {
        margin: 0;
        padding: 0;
        font-size: 2rem;
    }

    .rgpt-welcome p {
        margin-top: 0.35rem;
        opacity: 0.78;
        font-size: 1rem;
    }

    .rgpt-feature-card {
        padding: 0.75rem 0.9rem;
        border: 1px solid rgba(128,128,128,.22);
        border-radius: 12px;
        margin-bottom: .6rem;
    }

    """,
    unsafe_allow_html=True,
)


# ============================================================
# PUBLIC CREATOR ATTRIBUTION / SEO METADATA
# ============================================================

st.markdown(
    f"""
    <meta name="author" content="{escape(CREATOR_NAME)}">
    <meta name="creator" content="{escape(CREATOR_NAME)}">
    <meta name="description" content="RacharlaGPT, an AI learning and programming assistant created and developed by {escape(CREATOR_NAME)}. Supported by {escape(SUPPORTERS_NAME)}.">
    <script type="application/ld+json">
    {{
      "@context": "https://schema.org",
      "@type": "SoftwareApplication",
      "name": "RacharlaGPT",
      "applicationCategory": "EducationalApplication",
      "creator": {{"@type": "Person", "name": "{escape(CREATOR_NAME)}"}},
      "description": "AI learning and programming assistant"
    }}
    </script>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# SESSION PERSISTENCE MANAGEMENT (QUERY PARAMS + GOOGLE OAUTH)
# ============================================================

def restore_session():
    """Restore an existing session or complete the Google OAuth PKCE callback."""
    if "auth_user" in st.session_state and st.session_state.auth_user:
        return

    params = st.query_params
    if params.get("logged_out") == "1":
        return

    access_token = params.get("access_token")
    refresh_token = params.get("refresh_token")
    code = params.get("code")

    # Google/Supabase PKCE callback. The same Supabase client must be used
    # here as the one that created the OAuth URL so its code verifier remains
    # available for the one-time code exchange.
    if code:
        try:
            res = supabase.auth.exchange_code_for_session({"auth_code": code})
            if res.user and res.session:
                st.session_state.auth_user = res.user
                st.query_params["access_token"] = res.session.access_token
                st.query_params["refresh_token"] = res.session.refresh_token
                st.query_params.pop("code", None)
                st.query_params.pop("logged_out", None)
                st.session_state.pop("google_oauth_error", None)
                st.session_state.pop("google_oauth_url", None)
                st.rerun()
                return
        except Exception as exc:
            st.session_state.google_oauth_error = str(exc)

    # Restore via access/refresh tokens (used by email/password login and
    # by callbacks that return tokens directly).
    if access_token and refresh_token:
        try:
            res = supabase.auth.set_session(access_token, refresh_token)
            if res.user:
                st.session_state.auth_user = res.user
                st.query_params.pop("logged_out", None)
                return
        except Exception as exc:
            st.session_state.google_oauth_error = str(exc)

    # Fallback to any session already held by the Supabase client.
    try:
        session = supabase.auth.get_session()
        if session and session.user:
            st.session_state.auth_user = session.user
            if getattr(session, "access_token", None):
                st.query_params["access_token"] = session.access_token
            if getattr(session, "refresh_token", None):
                st.query_params["refresh_token"] = session.refresh_token
            st.query_params.pop("logged_out", None)
    except Exception:
        pass


restore_session()


# ============================================================
# AUTHENTICATION PAGE
# ============================================================

def show_auth():

    left, right = st.columns(
        [1.08, 0.92],
        gap="large",
    )

    # ========================================================
    # LEFT — BRANDING
    # ========================================================

    with left:

        st.markdown(
            '<div class="auth-left">',
            unsafe_allow_html=True,
        )

        if LOGO_PATH.exists():

            st.markdown(
                '<div class="auth-logo-wrap">',
                unsafe_allow_html=True,
            )

            st.image(
                str(LOGO_PATH),
                width=78,
            )

            st.markdown(
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div style="font-size:48px;margin-bottom:6px;">⚡</div>',
                unsafe_allow_html=True,
            )

        st.markdown(
            """
            <div class="auth-heading">
                RacharlaGPT
            </div>

            <div class="auth-description">
                A practical AI assistant for learning, coding,
                ideas, research, and everyday questions.
            </div>

            <div class="auth-badge-row">
                <span class="auth-badge">⚡ Fast AI</span>
                <span class="auth-badge">☁️ Saved Chats</span>
                <span class="auth-badge">🔐 Private Account</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        if BANNER_PATH.exists():

            st.markdown(
                '<div class="auth-banner-box">',
                unsafe_allow_html=True,
            )

            st.image(
                str(BANNER_PATH),
                width=590,
            )

            st.markdown(
                '</div>',
                unsafe_allow_html=True,
            )
        else:
            st.info(
                "ManaTechSaavy banner is not available in the deployed assets folder. "
                "Add assets/manatechsaavy_banner.jfif to the repository."
            )

        st.markdown(
            f'<div style="margin-top:10px;color:#7a8494;font-size:12px;">{escape(CREATOR_CREDIT)} • Supporters: {escape(SUPPORTERS_NAME)}</div>',
            unsafe_allow_html=True,
        )

        st.markdown(
            '</div>',
            unsafe_allow_html=True,
        )


    # ========================================================
    # RIGHT — LOGIN CARD
    # ========================================================

    with right:

        st.markdown(
            '<div class="auth-right">',
            unsafe_allow_html=True,
        )

        st.markdown(
            """
            <div class="auth-card-title">
                Welcome back 👋
            </div>

            <div class="auth-card-subtitle">
                Sign in to save and access your conversations permanently.
            </div>
            """,
            unsafe_allow_html=True,
        )

        # ----------------------------------------------------
        # ONE-CLICK GOOGLE / GMAIL SIGN-IN
        # ----------------------------------------------------
        # Streamlit renders custom HTML inside its app frame. The previous
        # version tried to navigate with JavaScript, but the browser can block
        # that navigation from the embedded frame.
        #
        # Instead, create the Supabase OAuth URL once and use a real HTML
        # link with target="_top". Clicking it performs a normal top-level
        # browser navigation to Google, which is the correct OAuth behavior.
        # The URL is stored in session state so we do not recreate the PKCE
        # flow on every Streamlit rerun.
        if "google_oauth_url" not in st.session_state:
            try:
                oauth_response = supabase.auth.sign_in_with_oauth(
                    {
                        "provider": "google",
                        "options": {
                            "redirect_to": APP_URL,
                        },
                    }
                )

                google_oauth_url = getattr(oauth_response, "url", None)

                if not google_oauth_url:
                    raise RuntimeError(
                        "Supabase did not return a Google OAuth authorization URL."
                    )

                st.session_state.google_oauth_url = google_oauth_url

            except Exception as exc:
                st.session_state.google_oauth_error = str(exc)

        google_oauth_url = st.session_state.get("google_oauth_url")

        if google_oauth_url:
            safe_oauth_url = escape(google_oauth_url, quote=True)
            st.markdown(
                f"""
                <a href="{safe_oauth_url}" target="_top" rel="noopener noreferrer"
                   style="display:flex;align-items:center;justify-content:center;
                          width:100%;box-sizing:border-box;padding:0.58rem 1rem;
                          border-radius:0.5rem;background:#ff4b4b;color:white;
                          text-decoration:none;font-weight:600;font-size:1rem;
                          border:1px solid #ff4b4b;cursor:pointer;">
                    🔵&nbsp;&nbsp;Continue with Google (Gmail)
                </a>
                """,
                unsafe_allow_html=True,
            )
        else:
            st.error("Google sign-in could not be started.")

        st.caption(
            "One click • Use your Google account • No RacharlaGPT password to remember"
        )

        if st.session_state.get("google_oauth_error"):
            st.error(
                "Google sign-in returned to RacharlaGPT, but the secure session "
                "exchange failed. Please start Google sign-in again."
            )
            with st.expander("Google sign-in diagnostic"):
                st.code(st.session_state.google_oauth_error)
            if st.button("🔄 Try Google sign-in again", use_container_width=True):
                st.session_state.pop("google_oauth_error", None)
                st.session_state.pop("google_oauth_url", None)
                st.rerun()

        st.markdown('<div class="divider-container">OR OTHER SIGN-IN OPTIONS</div>', unsafe_allow_html=True)

        # Keep email/password available as a secondary option, but do not
        # make new users deal with it unless they actually want it.
        with st.expander("Use email and password instead"):
            login_tab, signup_tab = st.tabs(
                [
                    "🔐 Sign In",
                    "✨ Create Account",
                ]
            )


            # ====================================================
            # SIGN IN
            # ====================================================

            with login_tab:

                with st.form("login_form"):

                    email = st.text_input(
                        "Email",
                        key="login_email",
                        placeholder="you@example.com",
                    )

                    password = st.text_input(
                        "Password",
                        type="password",
                        key="login_password",
                        placeholder="Enter your password",
                    )

                    submit = st.form_submit_button(
                        "Sign In",
                        type="primary",
                        use_container_width=True,
                    )


                if submit:

                    if not email.strip() or not password:

                        st.warning(
                            "Enter your email and password."
                        )

                    else:

                        try:

                            result = (
                                supabase.auth
                                .sign_in_with_password(
                                    {
                                        "email": email.strip(),
                                        "password": password,
                                    }
                                )
                            )

                            if result.user and result.session:

                                st.session_state.auth_user = (
                                    result.user
                                )

                                # Persist session details across reloads using Query Parameters
                                st.query_params["access_token"] = result.session.access_token
                                st.query_params["refresh_token"] = result.session.refresh_token
                                st.query_params.pop("logged_out", None)

                                st.rerun()

                            else:

                                st.error(
                                    "Sign in failed. "
                                    "Please check your email and password."
                                )

                        except Exception as exc:

                            st.error(
                                f"Sign in failed: {exc}"
                            )


            # ====================================================
            # CREATE ACCOUNT
            # ====================================================

            with signup_tab:

                with st.form("signup_form"):

                    email = st.text_input(
                        "Email",
                        key="signup_email",
                        placeholder="you@example.com",
                    )

                    password = st.text_input(
                        "Password",
                        type="password",
                        key="signup_password",
                        placeholder="Create a password",
                    )

                    confirm = st.text_input(
                        "Confirm password",
                        type="password",
                        key="signup_confirm",
                        placeholder="Repeat your password",
                    )

                    submit = st.form_submit_button(
                        "Create Account",
                        type="primary",
                        use_container_width=True,
                    )


                if submit:

                    if not email.strip() or not password:

                        st.warning(
                            "Enter your email and password."
                        )

                    elif password != confirm:

                        st.warning(
                            "Passwords do not match."
                        )

                    elif len(password) < 6:

                        st.warning(
                            "Password must be at least 6 characters."
                        )

                    else:

                        try:

                            result = supabase.auth.sign_up(
                                {
                                    "email": email.strip(),
                                    "password": password,
                                }
                            )

                            if result.session and result.user:

                                st.session_state.auth_user = (
                                    result.user
                                )

                                # Direct log-in persistence when email verification is disabled
                                st.query_params["access_token"] = result.session.access_token
                                st.query_params["refresh_token"] = result.session.refresh_token
                                st.query_params.pop("logged_out", None)

                                st.rerun()

                            else:

                                st.success(
                                    "Account created. "
                                    "You can now sign in."
                                )

                        except Exception as exc:

                            st.error(
                                f"Account creation failed: {exc}"
                            )


            st.markdown(
                """
                <div style="
                    text-align:center;
                    color:#8a94a6;
                    font-size:11px;
                    margin-top:11px;
                ">
                    🔒 Your chat history is securely stored in Supabase.
                </div>
                """,
                unsafe_allow_html=True,
            )

        st.markdown(
            '</div>',
            unsafe_allow_html=True,
        )


# ============================================================
# AUTH GATE
# ============================================================

if "auth_user" not in st.session_state or st.session_state.auth_user is None:

    show_auth()

    st.stop()


auth_user = st.session_state.auth_user

def _rgpt_display_name(user):
    metadata = getattr(user, "user_metadata", None) or {}
    name = metadata.get("full_name") or metadata.get("name") or metadata.get("user_name")
    if name:
        return str(name).strip()
    email = getattr(user, "email", None)
    return str(email).split("@")[0] if email else "Learner"

def _rgpt_avatar_url(user):
    metadata = getattr(user, "user_metadata", None) or {}
    url = metadata.get("avatar_url") or metadata.get("picture")
    return url if isinstance(url, str) and url.startswith(("https://", "http://")) else None

def render_rgpt_welcome(user):
    name = _rgpt_display_name(user)
    avatar = _rgpt_avatar_url(user)
    left, right = st.columns([5, 1])
    with left:
        st.markdown(
            f'<div class="rgpt-welcome"><div class="rgpt-kicker">WELCOME TO RACHARLAGPT</div>'
            f'<h1>Hi, {name} 👋</h1>'
            f'<p>Your learning, coding, data, AI, career and everyday assistant.</p></div>',
            unsafe_allow_html=True,
        )
    with right:
        local_avatar = st.session_state.get("profile_avatar_bytes")
        if local_avatar:
            st.image(local_avatar, width=76)
        elif avatar:
            try:
                st.image(avatar, width=76)
            except Exception:
                pass

def render_rgpt_getting_started():
    with st.expander("✨ New here? Start with RacharlaGPT", expanded=False):
        st.markdown(
            """
**📚 Learn** — Python, SQL, Excel, Power BI, JavaScript, Data Analytics, AI and more.

**🧪 Practice** — quizzes, coding challenges, SQL exercises and interview questions.

**🏗️ Build** — real-world projects, datasets, dashboards, portfolios and step-by-step guidance.

**🔍 Analyze** — screenshots, errors, charts, documents, resumes and study material.

**🎙️ Voice** — speak your question and continue from the transcription.

**💬 General Assistant** — ask everyday questions too.

**Try:** “Teach me SQL from beginner to advanced” • “Analyze this screenshot” • “Give me a project and evaluate my solution.”
"""
        )


# Enhanced post-login welcome UI
render_rgpt_welcome(auth_user)
render_rgpt_getting_started()
USER_ID = str(auth_user.id)


# ============================================================
# MODELS
# ============================================================

@st.cache_resource(show_spinner=False)
def get_models(api_key):

    primary = ChatGroq(
        model=PRIMARY_MODEL,
        temperature=0.7,
        api_key=api_key,
    )

    backup = ChatGroq(
        model=BACKUP_MODEL,
        temperature=0.7,
        api_key=api_key,
    )

    return primary, backup


primary_llm, backup_llm = get_models(
    GROQ_API_KEY
)


# ============================================================
# SESSION STATE
# ============================================================

if "system_prompt" not in st.session_state:
    st.session_state.system_prompt = DEFAULT_SYSTEM_PROMPT

if "chats" not in st.session_state:
    st.session_state.chats = {}

if "current_chat_id" not in st.session_state:
    st.session_state.current_chat_id = None

if "loaded_from_supabase" not in st.session_state:
    st.session_state.loaded_from_supabase = False

if "search" not in st.session_state:
    st.session_state.search = ""

if "selected_model" not in st.session_state:
    st.session_state.selected_model = "Auto (recommended)"

if "temperature" not in st.session_state:
    st.session_state.temperature = 0.7

if "course_focus" not in st.session_state:
    st.session_state.course_focus = "Auto Detect"

if "learning_mode" not in st.session_state:
    st.session_state.learning_mode = "Auto (best for the question)"

if "learner_level" not in st.session_state:
    st.session_state.learner_level = "Auto Detect"

if "answer_style" not in st.session_state:
    st.session_state.answer_style = "Balanced"


# ============================================================
# CHAT FUNCTIONS
# ============================================================

def blank_chat():

    now = datetime.now(timezone.utc).isoformat()

    return {
        "id": str(uuid.uuid4()),
        "user_id": USER_ID,
        "title": "New Chat",
        "messages": [],
        "created_at": now,
        "updated_at": now,
    }


def load_chats():

    result = (
        supabase
        .table("chats")
        .select("*")
        .eq("user_id", USER_ID)
        .order("updated_at", desc=True)
        .execute()
    )

    chats = {}

    for row in result.data or []:

        chats[str(row["id"])] = {
            "id": str(row["id"]),
            "user_id": str(row["user_id"]),
            "title": row.get("title") or "New Chat",
            "messages": row.get("messages") or [],
            "created_at": row.get("created_at"),
            "updated_at": row.get("updated_at"),
        }

    if not chats:

        chat = blank_chat()

        (
            supabase
            .table("chats")
            .insert(chat)
            .execute()
        )

        chats[chat["id"]] = chat

    return chats


def save_chat(chat):
    """Persist the complete chat and verify Supabase actually returned a row."""
    now = datetime.now(timezone.utc).isoformat()
    payload = {
        "id": str(chat["id"]),
        "user_id": USER_ID,
        "title": chat.get("title") or "New Chat",
        "messages": chat.get("messages") or [],
        "created_at": chat.get("created_at") or now,
        "updated_at": now,
    }
    result = (
        supabase
        .table("chats")
        .upsert(payload, on_conflict="id")
        .select("id,user_id,title,messages,created_at,updated_at")
        .execute()
    )
    if not result.data:
        raise RuntimeError("Supabase did not return the saved chat row.")
    saved = result.data[0]
    chat["id"] = str(saved.get("id") or chat["id"])
    chat["user_id"] = str(saved.get("user_id") or USER_ID)
    chat["title"] = saved.get("title") or chat.get("title") or "New Chat"
    chat["messages"] = saved.get("messages") or chat.get("messages") or []
    chat["created_at"] = saved.get("created_at") or chat.get("created_at") or now
    chat["updated_at"] = saved.get("updated_at") or now


def create_chat():
    chat = blank_chat()
    result = (
        supabase
        .table("chats")
        .insert(chat)
        .select("id,user_id,title,messages,created_at,updated_at")
        .execute()
    )
    if not result.data:
        raise RuntimeError("Supabase did not create the new chat.")
    row = result.data[0]
    chat["id"] = str(row.get("id") or chat["id"])
    chat["user_id"] = str(row.get("user_id") or USER_ID)
    chat["title"] = row.get("title") or chat["title"]
    chat["messages"] = row.get("messages") or []
    chat["created_at"] = row.get("created_at") or chat["created_at"]
    chat["updated_at"] = row.get("updated_at") or chat["updated_at"]
    st.session_state.chats[chat["id"]] = chat
    st.session_state.current_chat_id = chat["id"]


def delete_chat(chat_id):

    (
        supabase
        .table("chats")
        .delete()
        .eq("id", chat_id)
        .eq("user_id", USER_ID)
        .execute()
    )

    st.session_state.chats.pop(
        chat_id,
        None,
    )

    if not st.session_state.chats:

        create_chat()

    else:

        ordered = sorted(
            st.session_state.chats.values(),
            key=lambda x: x.get("updated_at") or "",
            reverse=True,
        )

        st.session_state.current_chat_id = ordered[0]["id"]


def delete_all():

    (
        supabase
        .table("chats")
        .delete()
        .eq("user_id", USER_ID)
        .execute()
    )

    st.session_state.chats = {}
    create_chat()


def title_for(text):

    text = " ".join(text.strip().split())

    if len(text) <= 34:
        return text

    return text[:34].rstrip() + "…"


def current_chat():

    if (
        st.session_state.current_chat_id
        not in st.session_state.chats
    ):
        create_chat()

    return st.session_state.chats[
        st.session_state.current_chat_id
    ]



# ============================================================
# ATTACHMENTS, IMAGES, AND VOICE
# ============================================================

VISION_MAX_IMAGES = 5
VISION_MAX_BYTES = 3_500_000


def _image_bytes_for_vision(uploaded_file):
    """Return a compact JPEG/PNG byte payload suitable for Groq vision."""
    raw = uploaded_file.getvalue()
    if len(raw) <= VISION_MAX_BYTES:
        return raw, uploaded_file.type or "image/jpeg"

    try:
        from PIL import Image

        image = Image.open(io.BytesIO(raw)).convert("RGB")
        image.thumbnail((2400, 2400))
        output = io.BytesIO()
        quality = 88
        while quality >= 55:
            output.seek(0)
            output.truncate(0)
            image.save(output, format="JPEG", quality=quality, optimize=True)
            data = output.getvalue()
            if len(data) <= VISION_MAX_BYTES:
                return data, "image/jpeg"
            quality -= 8
    except Exception:
        pass

    return raw, uploaded_file.type or "image/jpeg"


def analyze_image_attachment(uploaded_file, question):
    """Use a Groq vision model to inspect an attached screenshot/photo/image."""
    image_bytes, mime_type = _image_bytes_for_vision(uploaded_file)

    if len(image_bytes) > VISION_MAX_BYTES:
        raise ValueError(
            f"{uploaded_file.name} is too large for image analysis. "
            "Please upload a smaller screenshot/image."
        )

    encoded = base64.b64encode(image_bytes).decode("utf-8")
    if mime_type not in {"image/jpeg", "image/png", "image/webp", "image/gif"}:
        mime_type = "image/jpeg"

    client = Groq(api_key=GROQ_API_KEY)
    completion = client.chat.completions.create(
        model=VISION_MODEL,
        messages=[
            {
                "role": "system",
                "content": (
                    "You are the visual-analysis component of RacharlaGPT. "
                    "Inspect the supplied image carefully. Extract visible text, "
                    "code, errors, tables, charts, UI elements, and other relevant "
                    "details. Do not invent details that are not visible."
                ),
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": (
                            "Analyze this attachment for the user's question.\n\n"
                            f"USER QUESTION:\n{question}\n\n"
                            "Return useful factual visual context that another assistant "
                            "can use to answer the question. Preserve code and important "
                            "visible text accurately."
                        ),
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{mime_type};base64,{encoded}"
                        },
                    },
                ],
            },
        ],
        temperature=0,
        max_completion_tokens=4096,
    )
    return completion.choices[0].message.content or ""


def extract_text_attachment(uploaded_file):
    """Extract useful text from common document/data attachments."""
    name = uploaded_file.name.lower()
    data = uploaded_file.getvalue()

    if name.endswith((".txt", ".md", ".py", ".sql", ".js", ".json", ".csv", ".ts", ".html", ".css")):
        return data.decode("utf-8", errors="replace")[:120000]

    if name.endswith(".pdf"):
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(data))
            parts = []
            for page in reader.pages[:80]:
                parts.append(page.extract_text() or "")
            text = "\n\n".join(parts).strip()
            return text[:120000] if text else "[PDF contains no extractable text; it may be scanned/image-only.]"
        except ImportError:
            return "[PDF text extraction is unavailable because pypdf is not installed.]"
        except Exception as exc:
            return f"[Could not extract PDF text: {exc}]"

    if name.endswith(".docx"):
        try:
            from docx import Document

            doc = Document(io.BytesIO(data))
            text = "\n".join(p.text for p in doc.paragraphs)
            return text[:120000]
        except ImportError:
            return "[DOCX extraction is unavailable because python-docx is not installed.]"
        except Exception as exc:
            return f"[Could not extract DOCX text: {exc}]"

    if name.endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            parts = []
            for ws in wb.worksheets[:10]:
                parts.append(f"### Sheet: {ws.title}")
                for row in ws.iter_rows(max_row=300, values_only=True):
                    parts.append(" | ".join("" if v is None else str(v) for v in row))
            return "\n".join(parts)[:120000]
        except ImportError:
            return "[Excel extraction is unavailable because openpyxl is not installed.]"
        except Exception as exc:
            return f"[Could not extract Excel data: {exc}]"

    return (
        f"[Unsupported attachment type: {uploaded_file.name}. "
        "Please use an image, PDF, DOCX, TXT, CSV, XLSX, or common code/text file.]"
    )


def build_attachment_context(uploaded_files, question):
    """Turn current uploads into compact context for the normal RacharlaGPT answer."""
    if not uploaded_files:
        return ""

    image_files = [
        f for f in uploaded_files
        if (f.type or "").startswith("image/")
        or f.name.lower().endswith((".png", ".jpg", ".jpeg", ".webp", ".gif"))
    ]

    parts = []
    if image_files:
        for image_file in image_files[:VISION_MAX_IMAGES]:
            try:
                visual = analyze_image_attachment(image_file, question)
                parts.append(
                    f"ATTACHED IMAGE/SCREENSHOT: {image_file.name}\n{visual}"
                )
            except Exception as exc:
                parts.append(
                    f"ATTACHED IMAGE/SCREENSHOT: {image_file.name}\n"
                    f"[Image analysis failed: {exc}]"
                )

    for file in uploaded_files:
        if file in image_files[:VISION_MAX_IMAGES]:
            continue
        text = extract_text_attachment(file)
        parts.append(f"ATTACHED FILE: {file.name}\n{text}")

    if not parts:
        return ""

    return (
        "\n\nATTACHMENT CONTEXT — USE THIS WHEN RELEVANT:\n"
        + "\n\n---\n\n".join(parts)
    )


def transcribe_voice(audio_file):
    """Convert microphone audio to text using Groq Whisper."""
    client = Groq(api_key=GROQ_API_KEY)
    audio_bytes = audio_file.getvalue()

    result = client.audio.transcriptions.create(
        file=(audio_file.name or "voice.wav", audio_bytes),
        model=TRANSCRIPTION_MODEL,
        response_format="json",
        temperature=0,
    )
    return (result.text or "").strip()


# ============================================================
# AI RESPONSE
# ============================================================

def ask(chat, latest_model_context=""):

    recent = chat["messages"][-MAX_CONTEXT_MESSAGES:]

    course_context = (
        COURSE_INTELLIGENCE_PROMPT
        + "\nSELECTED COURSE FOCUS: " + st.session_state.course_focus
        + "\nSELECTED LEARNING MODE: " + st.session_state.learning_mode
        + "\nLEARNER LEVEL: " + st.session_state.learner_level
        + "\nANSWER STYLE: " + st.session_state.answer_style
    )

    system_content = (
        st.session_state.system_prompt
        + "\n\n"
        + course_context
        + "\n\n"
        + "REAL-TIME DATE CONTEXT:\n"
        + date_context()
    )

    msgs = [
        SystemMessage(content=system_content)
    ]

    for message in recent:

        if message["role"] == "user":

            user_content = message["content"]
            if message is recent[-1] and latest_model_context:
                user_content += latest_model_context
            elif message.get("model_context"):
                user_content += message.get("model_context", "")

            msgs.append(
                HumanMessage(
                    content=user_content
                )
            )

        else:

            msgs.append(
                AIMessage(
                    content=message["content"]
                )
            )

    selected = st.session_state.selected_model


    # ========================================================
    # PRIMARY ONLY
    # ========================================================

    if selected == PRIMARY_MODEL:

        try:

            llm = ChatGroq(
                model=PRIMARY_MODEL,
                temperature=st.session_state.temperature,
                api_key=GROQ_API_KEY,
            )

            return llm.invoke(msgs), "primary"

        except RateLimitError:

            llm = ChatGroq(
                model=BACKUP_MODEL,
                temperature=st.session_state.temperature,
                api_key=GROQ_API_KEY,
            )

            return llm.invoke(msgs), "backup"


    # ========================================================
    # BACKUP ONLY
    # ========================================================

    if selected == BACKUP_MODEL:

        llm = ChatGroq(
            model=BACKUP_MODEL,
            temperature=st.session_state.temperature,
            api_key=GROQ_API_KEY,
        )

        return llm.invoke(msgs), "backup"


    # ========================================================
    # AUTO
    # ========================================================

    try:

        llm = ChatGroq(
            model=PRIMARY_MODEL,
            temperature=st.session_state.temperature,
            api_key=GROQ_API_KEY,
        )

        return llm.invoke(msgs), "primary"

    except RateLimitError:

        llm = ChatGroq(
            model=BACKUP_MODEL,
            temperature=st.session_state.temperature,
            api_key=GROQ_API_KEY,
        )

        return llm.invoke(msgs), "backup"


# ============================================================
# LOAD SUPABASE CHATS
# ============================================================

if not st.session_state.loaded_from_supabase:

    try:

        st.session_state.chats = load_chats()

        ordered = sorted(
            st.session_state.chats.values(),
            key=lambda x: x.get("updated_at") or "",
            reverse=True,
        )

        if ordered:

            st.session_state.current_chat_id = ordered[0]["id"]

        st.session_state.loaded_from_supabase = True

    except Exception as exc:

        st.error("Could not load chats from Supabase.")
        st.code(str(exc))
        st.stop()


# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:

    with st.expander("👤 Profile photo", expanded=False):
        st.caption("Optional. Your Google avatar is used when available.")
        uploaded_avatar = st.file_uploader(
            "Choose a photo",
            type=["png", "jpg", "jpeg", "webp"],
            key="profile_avatar_upload",
        )
        if uploaded_avatar is not None:
            st.session_state["profile_avatar_bytes"] = uploaded_avatar.getvalue()
            st.session_state["profile_avatar_name"] = uploaded_avatar.name

    # --------------------------------------------------------
    # LOGO + BRAND
    # --------------------------------------------------------

    if LOGO_PATH.exists():

        st.image(
            str(LOGO_PATH),
            width=55,
        )

    st.markdown(
        """
        <div style="
            font-size:25px;
            font-weight:800;
            letter-spacing:-0.5px;
            margin-bottom:13px;
        ">
            ⚡ RacharlaGPT
        </div>
        """,
        unsafe_allow_html=True,
    )


    # --------------------------------------------------------
    # NEW CHAT
    # --------------------------------------------------------

    if st.button(
        "＋ New Chat",
        type="primary",
        use_container_width=True,
    ):

        create_chat()
        st.rerun()


    # --------------------------------------------------------
    # ACCOUNT
    # --------------------------------------------------------

    user_email = getattr(
        auth_user,
        "email",
        "User",
    )

    st.caption(
        f"Signed in: {user_email}"
    )


    # --------------------------------------------------------
    # SIGN OUT
    # --------------------------------------------------------

    if st.button(
        "🚪 Sign Out",
        use_container_width=True,
    ):

        try:
            supabase.auth.sign_out()
        except Exception:
            pass

        # Clear query parameter tokens on sign out
        st.query_params.clear()
        st.query_params["logged_out"] = "1"

        for key in [
            "auth_user",
            "google_oauth_url",
            "google_oauth_error",
            "chats",
            "current_chat_id",
            "loaded_from_supabase",
            "selected_model",
            "temperature",
            "course_focus",
            "learning_mode",
            "learner_level",
            "answer_style",
        ]:

            st.session_state.pop(
                key,
                None,
            )

        st.rerun()


    st.markdown(
        "### YOUR CHATS"
    )


    # --------------------------------------------------------
    # SEARCH
    # --------------------------------------------------------

    search = st.text_input(
        "Search",
        placeholder="🔎 Search your chats...",
        label_visibility="collapsed",
    )

    st.session_state.search = search


    # --------------------------------------------------------
    # RENAME
    # --------------------------------------------------------

    if (
        st.session_state.current_chat_id
        in st.session_state.chats
    ):

        with st.expander("✏️ Rename current chat"):

            rename_value = st.text_input(
                "Chat name",
                value=st.session_state.chats[
                    st.session_state.current_chat_id
                ]["title"],
                key="rename_chat_title",
            )

            if st.button(
                "Save name",
                use_container_width=True,
            ):

                rename_value = " ".join(
                    rename_value.strip().split()
                )

                if rename_value:

                    current = st.session_state.chats[
                        st.session_state.current_chat_id
                    ]

                    current["title"] = rename_value[:80]

                    save_chat(current)

                    st.rerun()


    # --------------------------------------------------------
    # CHAT LIST
    # --------------------------------------------------------

    ordered = sorted(
        st.session_state.chats.items(),
        key=lambda x: x[1].get("updated_at") or "",
        reverse=True,
    )

    for chat_id, item in ordered:

        q = search.lower().strip()

        matches = (
            not q
            or q in item["title"].lower()
            or any(
                q in str(
                    message.get("content", "")
                ).lower()
                for message in item["messages"]
            )
        )

        if not matches:
            continue

        c1, c2 = st.columns([4, 1])

        with c1:

            if st.button(
                f"💬 {item['title']}",
                key=f"open_{chat_id}",
                use_container_width=True,
            ):

                st.session_state.current_chat_id = chat_id
                st.rerun()

        with c2:

            if st.button(
                "🗑️",
                key=f"del_{chat_id}",
                use_container_width=True,
            ):

                delete_chat(chat_id)
                st.rerun()


    st.divider()


    # ========================================================
    # MANA TECH SAAVY LEARNING CENTER
    # ========================================================

    with st.expander("🎓 Learning Center", expanded=True):

        st.caption(
            "Choose how RacharlaGPT should teach. "
            "You can change these anytime."
        )

        st.session_state.course_focus = st.selectbox(
            "Course focus",
            COURSE_OPTIONS,
            index=COURSE_OPTIONS.index(
                st.session_state.course_focus
            ),
            help=(
                "Auto Detect lets RacharlaGPT identify the course "
                "from your question."
            ),
        )

        st.session_state.learning_mode = st.selectbox(
            "Learning mode",
            LEARNING_MODES,
            index=LEARNING_MODES.index(
                st.session_state.learning_mode
            ),
        )

        st.session_state.learner_level = st.selectbox(
            "Your level",
            LEVELS,
            index=LEVELS.index(
                st.session_state.learner_level
            ),
        )

        st.session_state.answer_style = st.selectbox(
            "Answer style",
            ANSWER_STYLES,
            index=ANSWER_STYLES.index(
                st.session_state.answer_style
            ),
        )

        st.info(
            "💡 Tip: Ask for a lesson, practice questions, "
            "a project, interview preparation, or debugging help."
        )


    st.divider()


    # ========================================================
    # DOWNLOAD
    # ========================================================

    current_export = st.session_state.chats.get(
        st.session_state.current_chat_id
    )

    if current_export:

        title = (
            current_export["title"]
            or "RacharlaGPT Chat"
        )

        txt_lines = [
            f"{title}\n",
            "RacharlaGPT Conversation\n",
            "Created and developed by Racharla Saikrishna\n",
            "Supporters: RSKT\n",
            "=" * 50 + "\n",
        ]

        for message in current_export.get("messages", []):

            role = (
                "You"
                if message.get("role") == "user"
                else "RacharlaGPT"
            )

            txt_lines.append(f"\n{role}\n")
            txt_lines.append("-" * 30 + "\n")
            txt_lines.append(
                str(message.get("content", ""))
            )
            txt_lines.append("\n")

        text_download = "".join(txt_lines)


        # ----------------------------------------------------
        # SELF-CONTAINED HTML EXPORT
        # ----------------------------------------------------

        html_parts = [
            "<!DOCTYPE html>",
            "<html>",
            "<head>",
            '<meta charset="UTF-8">',
            '<meta name="viewport" content="width=device-width, initial-scale=1.0">',
            f"<title>{escape(title)} — RacharlaGPT</title>",
            f'<meta name="author" content="{escape(CREATOR_NAME)}">',
            f'<meta name="description" content="RacharlaGPT conversation — created and developed by {escape(CREATOR_NAME)}; supporters {escape(SUPPORTERS_NAME)}.">',
            """
            <style>
                body {
                    font-family: Arial, sans-serif;
                    max-width: 900px;
                    margin: auto;
                    padding: 20px;
                    line-height: 1.6;
                    background: #ffffff;
                    color: #222222;
                }

                h1 {
                    font-size: 26px;
                }

                .message {
                    border: 1px solid #dddddd;
                    border-radius: 12px;
                    padding: 15px;
                    margin: 15px 0;
                }

                .role {
                    font-weight: bold;
                    margin-bottom: 8px;
                }

                .user {
                    background: #f5f5f5;
                }

                .assistant {
                    background: #ffffff;
                }

                .content {
                    white-space: pre-wrap;
                    word-wrap: break-word;
                }

                .footer {
                    margin-top: 30px;
                    color: #777777;
                    font-size: 12px;
                    text-align: center;
                }
            </style>
            """,
            "</head>",
            "<body>",
            f"<h1>{escape(title)}</h1>",
        ]

        for message in current_export.get("messages", []):

            is_user = (
                message.get("role") == "user"
            )

            role = (
                "You"
                if is_user
                else "RacharlaGPT"
            )

            css_class = (
                "user"
                if is_user
                else "assistant"
            )

            content = escape(
                str(
                    message.get(
                        "content",
                        "",
                    )
                )
            )

            html_parts.append(
                f"""
                <div class="message {css_class}">
                    <div class="role">
                        {role}
                    </div>

                    <div class="content">
                        {content}
                    </div>
                </div>
                """
            )

        html_parts.extend(
            [
                """
                <div class="footer">
                    RacharlaGPT • Created and developed by Racharla Saikrishna • Supporters: RSKT • Powered by Groq
                </div>
                """,
                "</body>",
                "</html>",
            ]
        )

        html_download = "".join(html_parts)

        st.markdown("### 📥 Download")

        st.download_button(
            "📄 Download TXT",
            data=text_download,
            file_name=f"{title[:45]}.txt",
            mime="text/plain",
            use_container_width=True,
        )

        st.download_button(
            "🌐 Download HTML",
            data=html_download,
            file_name=f"{title[:45]}.html",
            mime="text/html",
            use_container_width=True,
        )

        st.caption(
            "TXT is the most universal format. "
            "HTML keeps headings and conversation formatting."
        )


    # --------------------------------------------------------
    # DELETE ALL
    # --------------------------------------------------------

    if st.button(
        "🗑️ Delete All Chats",
        use_container_width=True,
    ):

        delete_all()
        st.rerun()


    st.divider()


    # ========================================================
    # AI SETTINGS
    # ========================================================

    with st.expander(
        "⚙️ AI Settings",
        expanded=False,
    ):

        st.session_state.selected_model = st.selectbox(
            "Model",
            MODEL_OPTIONS,
            index=MODEL_OPTIONS.index(
                st.session_state.selected_model
            ),
        )

        st.session_state.temperature = st.slider(
            "Creativity",
            min_value=0.0,
            max_value=1.2,
            value=float(
                st.session_state.temperature
            ),
            step=0.1,
            help=(
                "Lower = more focused. "
                "Higher = more creative."
            ),
        )

        st.session_state.system_prompt = st.text_area(
            "AI System Prompt",
            value=st.session_state.system_prompt,
            height=170,
        )

        if st.button(
            "↩️ Reset AI Settings",
            use_container_width=True,
        ):

            st.session_state.selected_model = (
                "Auto (recommended)"
            )

            st.session_state.temperature = 0.7

            st.session_state.system_prompt = (
                DEFAULT_SYSTEM_PROMPT
            )

            st.session_state.course_focus = "Auto Detect"
            st.session_state.learning_mode = "Auto (best for the question)"
            st.session_state.learner_level = "Auto Detect"
            st.session_state.answer_style = "Balanced"

            st.rerun()


    st.caption(
        f"Primary: {PRIMARY_MODEL}"
    )

    st.caption(
        f"Fallback: {BACKUP_MODEL}"
    )

    st.caption(
        "☁️ Chats stored in Supabase"
    )


# ============================================================
# MAIN APPLICATION BRANDING — UI ONLY
# ============================================================

brand_col1, brand_col2 = st.columns(
    [0.75, 11.25],
    gap="small",
)

with brand_col1:

    if LOGO_PATH.exists():

        st.image(
            str(LOGO_PATH),
            width=48,
        )

    else:

        st.markdown("⚡")


with brand_col2:

    st.markdown(
        """
        <div class="main-brand-name">
            RacharlaGPT
        </div>

        <div class="main-brand-subtitle">
            Fast, intelligent AI conversations powered by Groq
        </div>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# CURRENT CHAT
# ============================================================

chat = current_chat()


# ============================================================
# EMPTY CHAT
# ============================================================

if not chat["messages"]:

    st.markdown(
        """
        <div style="text-align:center; padding: 8vh 12px 4vh 12px;">
            <h1 style="font-size:31px; margin-bottom:8px;">
                Welcome to RacharlaGPT
            </h1>
            <p style="color:#707b8c; line-height:1.65; font-size:15px; margin:0 auto; max-width:680px;">
                Your conversations are saved to your account.<br>
                Ask questions, write code, brainstorm ideas,<br>
                learn something new, or simply have a conversation.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# CHAT HISTORY
# ============================================================

for message in chat["messages"]:

    with st.chat_message(message["role"]):

        st.markdown(
            message["content"]
        )


# ============================================================
# MAIN EXPORT — ALWAYS AVAILABLE FOR THE CURRENT CHAT
# ============================================================

with st.expander("📥 Download current chat", expanded=False):
    export_chat = st.session_state.chats.get(st.session_state.current_chat_id)
    if export_chat:
        export_title = export_chat.get("title") or "RacharlaGPT Chat"
        export_txt = [
            f"{export_title}\n",
            "RacharlaGPT Conversation\n",
            "Created and developed by Racharla Saikrishna\n",
            "Supporters: RSKT\n",
            "=" * 50 + "\n",
        ]
        for export_message in export_chat.get("messages", []):
            export_role = "You" if export_message.get("role") == "user" else "RacharlaGPT"
            export_txt.append(f"\n{export_role}\n")
            export_txt.append("-" * 30 + "\n")
            export_txt.append(str(export_message.get("content", "")))
            export_txt.append("\n")
        export_txt_data = "".join(export_txt)
        export_html = (
            "<!DOCTYPE html><html><head><meta charset=\"UTF-8\">"
            f"<meta name=\"author\" content=\"{escape(CREATOR_NAME)}\">"
            f"<title>{escape(export_title)} — RacharlaGPT</title></head><body>"
            f"<h1>{escape(export_title)}</h1>"
            f"<p><strong>Created and developed by {escape(CREATOR_NAME)}</strong> • Supporters: {escape(SUPPORTERS_NAME)}</p>"
            + "".join(
                f"<section><h3>{'You' if m.get('role') == 'user' else 'RacharlaGPT'}</h3>"
                f"<div style=\"white-space:pre-wrap;word-wrap:break-word;\">{escape(str(m.get('content','')))}</div></section>"
                for m in export_chat.get("messages", [])
            )
            + "</body></html>"
        )
        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "📄 Download TXT",
                data=export_txt_data,
                file_name=f"{export_title[:45]}.txt",
                mime="text/plain",
                use_container_width=True,
                key="main_download_txt",
            )
        with d2:
            st.download_button(
                "🌐 Download HTML",
                data=export_html,
                file_name=f"{export_title[:45]}.html",
                mime="text/html",
                use_container_width=True,
                key="main_download_html",
            )
    else:
        st.caption("Start a chat to enable downloads.")


# ============================================================
# ATTACHMENTS
# ============================================================

with st.expander("📎 Attach image, screenshot, PDF, document, or data", expanded=False):
    uploaded_files = st.file_uploader(
        "Attach files for RacharlaGPT to read or analyze",
        type=[
            "png", "jpg", "jpeg", "webp", "gif",
            "pdf", "docx", "txt", "md", "csv",
            "xlsx", "xlsm", "py", "sql", "js", "ts", "json", "html", "css",
        ],
        accept_multiple_files=True,
        key="chat_attachments",
        help=(
            "Attach screenshots, screen chats, images, PDFs, documents, spreadsheets, "
            "or code. Ask your question in the chat box and RacharlaGPT will use the "
            "attachment as context."
        ),
    )

if uploaded_files:
    st.caption(
        "Attached: " + ", ".join(f.name for f in uploaded_files[:8])
        + (" …" if len(uploaded_files) > 8 else "")
    )


# ============================================================
# VOICE INPUT
# ============================================================

with st.expander("🎙️ Ask by voice", expanded=False):
    voice_audio = st.audio_input(
        "Record your question",
        sample_rate=16000,
        key="voice_question_input",
        help="Speak your question. RacharlaGPT will convert it to text and use it as your search/question.",
    )

    if voice_audio:
        voice_bytes = voice_audio.getvalue()
        import hashlib

        voice_hash = hashlib.sha256(voice_bytes).hexdigest()

        if st.session_state.get("voice_audio_hash") != voice_hash:
            with st.spinner("Converting voice to text…"):
                try:
                    transcript = transcribe_voice(voice_audio)
                    st.session_state.voice_audio_hash = voice_hash
                    st.session_state.voice_transcript = transcript
                except Exception as exc:
                    st.session_state.voice_audio_hash = voice_hash
                    st.session_state.voice_transcript = ""
                    st.session_state.voice_error = str(exc)

        if st.session_state.get("voice_transcript"):
            st.text_area(
                "Transcribed question",
                value=st.session_state.voice_transcript,
                height=100,
                key="voice_transcript_preview",
            )
            send_voice = st.button(
                "🎙️ Send transcribed question",
                use_container_width=True,
                key="send_voice_question",
            )
        else:
            send_voice = False
            if st.session_state.get("voice_error"):
                st.error(
                    "Voice transcription failed. "
                    + st.session_state.voice_error
                )
    else:
        send_voice = False


# ============================================================
# CHAT INPUT
# ============================================================

typed_input = st.chat_input(
    "Message RacharlaGPT..."
)

user_input = typed_input
if send_voice and st.session_state.get("voice_transcript"):
    user_input = st.session_state.voice_transcript


def process_user_message(user_input, attachment_context=""):
    user_input = user_input.strip()
    if not user_input:
        return

    chat = current_chat()

    # Keep the visible user message clean while giving the model the attachment context.
    model_user_content = user_input
    if attachment_context:
        model_user_content += attachment_context

    chat["messages"].append(
        {
            "role": "user",
            "content": user_input,
            "model_context": attachment_context,
        }
    )

    if chat["title"] == "New Chat":
        chat["title"] = title_for(user_input)

    try:
        save_chat(chat)
    except Exception as save_exc:
        st.warning(
            f"Your message is visible, but could not be saved to Supabase: {save_exc}"
        )

    with st.chat_message("user"):
        st.markdown(user_input)

        if attachment_context:
            st.caption("📎 Attachment context included")

    try:
        with st.chat_message("assistant"):
            with st.spinner("RacharlaGPT is thinking…"):
                response, used = ask(chat, latest_model_context=attachment_context)

            answer = response.content
            if not isinstance(answer, str):
                answer = str(answer)

            st.markdown(answer)

            if used == "backup":
                st.caption(
                    "Primary model was rate-limited; "
                    f"answered with {BACKUP_MODEL}."
                )

        chat["messages"].append(
            {
                "role": "assistant",
                "content": answer,
            }
        )

        save_chat(chat)
        st.rerun()

    except RateLimitError:
        with st.chat_message("assistant"):
            st.warning(
                "Groq rate limit reached. Please wait for the quota to reset "
                "and try again."
            )
        try:
            save_chat(chat)
        except Exception:
            pass

    except Exception as exc:
        with st.chat_message("assistant"):
            st.error(f"RacharlaGPT could not complete the request: {exc}")
        try:
            save_chat(chat)
        except Exception:
            pass


if user_input:
    attachment_context = build_attachment_context(
        uploaded_files if "uploaded_files" in locals() else [],
        user_input,
    )

    if send_voice:
        st.session_state.voice_transcript = ""
        st.session_state.voice_error = ""

    process_user_message(
        user_input,
        attachment_context=attachment_context,
    )


if user_input:

    user_input = user_input.strip()

    if user_input:

        # ----------------------------------------------------
        # Save user message
        # ----------------------------------------------------

        chat["messages"].append(
            {
                "role": "user",
                "content": user_input,
            }
        )


        # ----------------------------------------------------
        # Automatic title
        # ----------------------------------------------------

        if chat["title"] == "New Chat":

            chat["title"] = title_for(
                user_input
            )


        # ----------------------------------------------------
        # Persist the user message immediately so it is not lost if Groq fails, rate-limits, or the browser is refreshed.
        try:
            save_chat(chat)
        except Exception as save_exc:
            st.warning(f"Your message is visible, but could not be saved to Supabase: {save_exc}")

        # Display user message
        # ----------------------------------------------------

        with st.chat_message("user"):

            st.markdown(
                user_input
            )


        # ----------------------------------------------------
        # Generate answer
        # ----------------------------------------------------

        try:

            with st.chat_message("assistant"):

                with st.spinner(
                    "RacharlaGPT is thinking…"
                ):

                    response, used = ask(chat)

                answer = response.content

                if not isinstance(answer, str):
                    answer = str(answer)

                st.markdown(answer)

                if used == "backup":

                    st.caption(
                        "Primary model was rate-limited; "
                        f"answered with {BACKUP_MODEL}."
                    )


            # ------------------------------------------------
            # Save assistant response
            # ------------------------------------------------

            chat["messages"].append(
                {
                    "role": "assistant",
                    "content": answer,
                }
            )

            save_chat(chat)
            st.rerun()


        except RateLimitError:

            with st.chat_message("assistant"):

                st.warning(
                    "Groq rate limit reached. "
                    "Please wait for the quota to reset "
                    "and try again."
                )

            try:
                save_chat(chat)
            except Exception:
                pass


        except Exception as exc:

            with st.chat_message("assistant"):

                st.error(
                    "Something went wrong while "
                    "contacting Groq. Please try again."
                )

            try:
                save_chat(chat)
            except Exception:
                pass

            print(
                "RacharlaGPT error:",
                type(exc).__name__,
                str(exc),
            )


# ============================================================
# FOOTER
# ============================================================

st.markdown(
    """
    <div style="
        text-align:center;
        color:#9aa0ad;
        font-size:11px;
        margin-top:15px;
        padding-bottom:10px;
    ">
        RacharlaGPT • Created and developed by Racharla Saikrishna • Supporters: RSKT • Powered by Groq • Chats stored in Supabase
    </div>
    """,
    unsafe_allow_html=True,
)
